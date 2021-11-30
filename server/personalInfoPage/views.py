from django.shortcuts import render
from NutriliteSearchPage.models import fileDataInfo,mainClassInfo,secClassInfo,personalFileData,personalExchangeFileLog
from userlogin.models import UserAccountInfo,UserAccountChainYenInfo, UserAccountAmwayInfo, registerDDandDimInfo
from NutriliteSearchPage.utils.page import Pagination
from pointManage.models import pointHistory
from django.contrib.auth.hashers import make_password,check_password
from django.http import HttpResponse
from managerPage.models import rewardReport
from django.core import serializers
from django.contrib.auth.decorators import permission_required
from openpyxl import load_workbook
import json
import datetime

# Create your views here.

def personalInfoHomePage(request,selectTag):
    if selectTag == "總部會議":
        selectTag += "/活動"


    try:
        userAccount = UserAccountInfo.objects.get(username=request.user)
        dataPermissionsLevel = userAccount.dataPermissionsLevel
        if selectTag == "全部":
            personalFile = userAccount.personalfiledata_set.filter(fileDataID__permissionsLevel__lte=dataPermissionsLevel,
                                                                   fileDataID__visible=1).order_by('exchangeDate')
        else:
            personalFile = userAccount.personalfiledata_set.filter(fileDataID__permissionsLevel__lte=dataPermissionsLevel,
                                                                   fileDataID__visible=1,
                                                                   fileDataID__mainClass__mainClassName=selectTag).order_by(
                'exchangeDate')
    except:
        dataPermissionsLevel = -1
    # fileDatas = fileDataInfo.objects.filter(mainClass=mainClassInfo.objects.get(mainClassName=selectTag).id,
    #                                         secClass=secClassInfo.objects.get(secClassName=selectTag).id,
    #                                         visible=1,
    #                                         permissionsLevel__lte=dataPermissionsLevel).order_by('occurrenceDate')
    fileDatas = personalFile
    # 總頁數
    page_count = fileDatas.count()
    pageisNotNull = True
    if page_count == 0:
        pageisNotNull = False
    # 當前頁
    current_page_num = request.GET.get("page")
    pagination = Pagination(current_page_num, page_count, request, per_page_num=10)
    # 處理之後的資料
    fileDatas = fileDatas[pagination.start:pagination.end]
    selectPage = "瀏覽曾兌換資料"
    content = {
        "fileDatas": fileDatas, "pagination": pagination, }


    return render(request, 'personalInfoPages/personalInfoHomePage.html', locals())


def personalInfoPointPage(request):

    username = request.user.username
    userAccountInfo = UserAccountInfo.objects.get(username = username)

    try:
        #轉讓: 白金小組 & 鑽石小組
        myTeam = []
        userAccountAmway = UserAccountAmwayInfo.objects.get(UserAccountInfo = userAccountInfo)
        awardRank = userAccountAmway.amwayAward.rank
        amNumber  = userAccountAmway.amwayNumber

        #白金
        if awardRank >= 15 and awardRank < 60:
            myDDInfo = registerDDandDimInfo.objects.get(amwayNumber = amNumber)
            accountIDList = UserAccountAmwayInfo.objects.filter(amwayDD = myDDInfo).values('UserAccountInfo')
            myTeam = UserAccountInfo.objects.filter(id__in=accountIDList).exclude(username = request.user.username)
        #鑽石
        elif awardRank >= 60:
            myDDInfo = registerDDandDimInfo.objects.get(amwayNumber = amNumber)
            accountIDList = UserAccountAmwayInfo.objects.filter(amwayDD = myDDInfo).values('UserAccountInfo')
            myTeam = UserAccountInfo.objects.filter(id__in=accountIDList).exclude(username = request.user.username)
            myTeamDD = registerDDandDimInfo.objects.filter(amwayDiamond = str(amNumber))
            myTeam = list(myTeam)
            for myDD in myTeamDD:
                accountIDList = UserAccountAmwayInfo.objects.filter(amwayDD = myDD).values('UserAccountInfo')
                if accountIDList:
                    ddTeam = UserAccountInfo.objects.filter(id__in = accountIDList).exclude(username = username)
                    for item in ddTeam:
                        if item not in myTeam:
                            myTeam.append(item)
        else:
            myTeam = []
        selectPage = "管理個人點數"
    except:
        # 此直銷權在資料庫沒有註冊為 DD
        myTeam = []
    return render(request, 'personalInfoPages/personalInfoPointPage.html', locals())



def changePasswordPage(request):

    return render(request, 'personalInfoPages/changePasswordPage.html', locals())


def changePasswordOption(request):
    password1 = request.POST.get("id_password1")
    password2 = request.POST.get("id_password2")
    password_old = request.POST.get("id_password_old")

    if password1 == password2:
        userAccountInfo = UserAccountInfo.objects.get(username=request.user.username)

        if check_password(password_old,userAccountInfo.password):
            userAccountInfo.password = make_password(password1)

            userAccountInfo.save()
            message = "成功修改密碼，五秒後跳轉到登入畫面"
        else:
            message = "修改失敗，舊密碼錯誤，五秒後跳轉到登入畫面"
    else:
        message = "修改失敗，發生錯誤，五秒後跳轉到登入畫面"

    # make_password(id_password1)
    #

    return render(request, 'personalInfoPages/changePasswordOptionResult.html', locals())

@permission_required('userlogin.rewardReport', login_url='/accounts/userlogin/')
def personalRewardReport(request):
    selectPage = "獎賞回報"
    return render(request, 'personalInfoPages/personalInfoRewadReportPage.html', locals())


@permission_required('userlogin.rewardReport', login_url='/accounts/userlogin/')
def getSelfRewardReport(request):
    res = HttpResponse()
    rewardAll = rewardReport.objects.filter(reporter = request.user)
    resContent = []
    for reward in rewardAll:
        rewardList = json.loads(reward.rewardList)
        rewardName = ""
        for user in rewardList:
            if rewardName == "":
                rewardName = user["user"]
            else:
                rewardName = rewardName + "," + user["user"]
            
        handleDate = reward.handleDate
        if not handleDate:
            handleDate = ""
        resTmp = {
            "reporter":   reward.reporter.user,
            "reason": reward.reason,
            "recordDate": str(reward.recordDate),
            "rewardList": rewardName,
            "status": reward.status,
            "handleDate": str(handleDate)
        }
        resContent.append(resTmp)
    res.status_code = 200
    res.content =  json.dumps(resContent)
    return res

@permission_required('userlogin.rewardReport', login_url='/accounts/userlogin/')
def rewardReportByExcel(request):
    res = HttpResponse()
    try:
        excelFile = ""
        if "excelFile" in request.FILES:
            excelFile = request.FILES['excelFile']
        
        reason = ""
        if "reason" in request.POST:
            reason = request.POST['reason']
        
        if excelFile == "" or reason == "" or \
            (reason == True and reason == False):
            res.content = "無提供檔案或是獎賞緣由"
            res.status_code = 400
            return res

        if not request.user.has_perm('userlogin.rewardReport'):
            res.content = "Permission not enough"
            res.status_code = 403
            return res

        wb = load_workbook(excelFile)
        sheet = wb.active
        resContent = []
        errorList  = []
        reporter = request.user
        for i in range(2,sheet.max_row+1):
            if not sheet['B' + str(i)].value:
                break
            errorTmp = {}
            resTmp = {}
            resTmp["user"] = str(sheet["A" + str(i)].value).strip(" ")
            amwayNumber = str(sheet['B' + str(i)].value)
            id4 = str(sheet['C' + str(i)].value)
            username = amwayNumber + id4
            resTmp["amwayNumber"] = amwayNumber
            resTmp["id4"] = id4
            resContent.append(resTmp)
            r = UserAccountInfo.objects.filter(username =  username)
            if not r or r.get().user != resTmp["user"]:
                errorTmp["user"] = resTmp["user"]
                errorTmp["amwayNumber"] = amwayNumber
                errorTmp["id4"] = id4
                errorList.append(errorTmp)
        if len(errorList) > 0:
            res.status_code = 404
            res.content = json.dumps(errorList)
            return res
        if len(resContent) == 0:
            res.status_code = 400
            res.content = "Excel 內無名單"
            return res
        rReport = rewardReport(reporter = reporter, reason = reason, 
                                recordDate = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                rewardList = json.dumps(resContent),
                                status = "waiting")
        rReport.save()
        res.status_code = 200
        res.content = json.dumps(resContent)
    except:
        res.status_code = 503
    
    return res