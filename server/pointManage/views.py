from django.shortcuts import render
from userlogin.models import UserAccountInfo, amwayAwardInfo, registerDDandDimInfo
from django.contrib.auth.decorators import permission_required
from userlogin.models import UserAccountInfo, UserAccountChainYenInfo, UserAccountAmwayInfo, chainYenJobTitleInfo
from django.db.models import F, Value
from django.http import HttpResponse
from pointManage.models import pointHistory
from django.core import serializers
import datetime, json
from openpyxl import load_workbook
# Create your views here.

@permission_required('userlogin.seeManagerPointPage', login_url='/accounts/userlogin/')
def addPoint(request):
    res = HttpResponse()
    try:
        userAccountInfo = UserAccountInfo.objects.get(username = request.POST['username'])
        accountChainyen = UserAccountChainYenInfo.objects.get(UserAccountInfo = userAccountInfo)
        point = accountChainyen.point + 10
        accountChainyen.point = F('point') + 10
        accountChainyen.save()

        res.status_code = 200
        res.content = point

        modifier = request.user.user
        pHistory = pointHistory(UserAccountInfo = userAccountInfo, modifier = modifier,
                                recordDate = datetime.datetime.now(), reason = '管理者加點',
                                addPoint = "+10", reducePoint = "", transferPoint = "",
                                resultPoint = point)
        pHistory.save()
    except:
        res.status_code = 503

    return res

@permission_required('userlogin.seeManagerPointPage', login_url='/accounts/userlogin/')
def addPointByAll(request):
    res = HttpResponse()
    try:
        point = 0
        try:
            point = int(request.POST["point"])
        except:
            point = 0
        if point == 0:
            res.status_code = 200
            return res

        userAccountChainyenList = UserAccountChainYenInfo.objects.all()
        userAccountChainyenList.update(point=F('point') + point)

        res.status_code = 200

        modifier = request.user.user
        for userAccountChainyen in userAccountChainyenList:
            userAccount = userAccountChainyen.UserAccountInfo
            resultPoint = userAccountChainyen.point
            pHistory = pointHistory(UserAccountInfo = userAccount, modifier = modifier,
                                    recordDate = datetime.datetime.now(), reason = '管理者加點',
                                    addPoint = "+" + str(point), reducePoint = "", transferPoint = "",
                                    resultPoint = resultPoint)
            pHistory.save()

    except:
        res.status_code = 404

    return res

@permission_required('userlogin.seeManagerPointPage', login_url='/accounts/userlogin/')
def addReducePointByUser(request):
    res = HttpResponse()
    try:
        point = 0
        try:
            point = int(request.POST["point"])
        except:
            point = 0
        if point == 0:
            res.status_code = 200
            return res

        username = ""
        try:
            username = request.POST["username"]
        except:
            username = ""
        if username == 0:
            res.status_code = 200
            return res

        userAccountInfo = UserAccountInfo.objects.get(username = username)
        userAccountChainyenList = UserAccountChainYenInfo.objects.filter(UserAccountInfo=userAccountInfo)
        userAccountChainyenList.update(point=F('point') + point)

        res.status_code = 200

        modifier = request.user.user
        if point > 0:
            reason = "管理者增點"
            addPoint = "+" + str(point)
            reducePoint = ""
        else:
            reason = "管理者扣點"
            addPoint = ""
            reducePoint = str(point)

        userAccount = userAccountChainyenList.get().UserAccountInfo
        resultPoint = userAccountChainyenList.get().point
        pHistory = pointHistory(UserAccountInfo = userAccount, modifier = modifier,
                                recordDate = datetime.datetime.now(), reason = reason,
                                addPoint = addPoint, reducePoint = reducePoint, transferPoint = "",
                                resultPoint = resultPoint)
        pHistory.save()
        res.content = resultPoint

    except:
        res.status_code = 404

    return res

@permission_required('userlogin.seeManagerPointPage', login_url='/accounts/userlogin/')
def addPointByAmwayAward(request):
    res = HttpResponse()
    try:
        amwayAwardStart = ""
        if "amwayAwardStart" in request.POST:
            amwayAwardStart = request.POST['amwayAwardStart']
        amwayAwardEnd   = ""
        if "amwayAwardEnd" in request.POST:
            amwayAwardEnd = request.POST['amwayAwardEnd']
        
        if amwayAwardStart:
            try:
                amwayAwardStart = amwayAwardInfo.objects.get(amwayAward=amwayAwardStart)
            except:
                amwayAwardStart = ""
        if amwayAwardEnd:
            try:
                amwayAwardEnd = amwayAwardInfo.objects.get(amwayAward=amwayAwardEnd)
            except:
                amwayAwardEnd = ""
        point = 0
        try:
            point = int(request.POST["point"])
        except:
            point = 0
        if point == 0:
            res.status_code = 200
            return res
        if amwayAwardEnd and amwayAwardStart:
            accountIDList = UserAccountAmwayInfo.objects.filter(amwayAward__rank__range = [amwayAwardStart.rank, amwayAwardEnd.rank]).values('UserAccountInfo')
            accountInfoList = UserAccountInfo.objects.filter(id__in=accountIDList)
            userAccountChainyenList = UserAccountChainYenInfo.objects.filter(UserAccountInfo__in=accountInfoList)
            userAccountChainyenList.update(point=F('point') + point)
        else:
            res.status_code = 404
            return res

        res.status_code = 200

        modifier = request.user.user
        for userAccountChainyen in userAccountChainyenList:
            userAccount = userAccountChainyen.UserAccountInfo
            resultPoint = userAccountChainyen.point
            pHistory = pointHistory(UserAccountInfo = userAccount, modifier = modifier,
                                    recordDate = datetime.datetime.now(), reason = '管理者加點',
                                    addPoint = "+" + str(point), reducePoint = "", transferPoint = "",
                                    resultPoint = resultPoint)
            pHistory.save()

    except:
        res.status_code = 404

    return res

@permission_required('userlogin.seeManagerPointPage', login_url='/accounts/userlogin/')
def addPointByJobTitle(request):
    res = HttpResponse()
    try:
        jobTitle = ""
        if "jobTitle" in request.POST:
            jobTitle = request.POST['jobTitle']
        if jobTitle:
            try:
                jobTitle = chainYenJobTitleInfo.objects.get(jobTitle = jobTitle)
            except:
                jobTitle = ""
        
        point = 0
        try:
            point = int(request.POST["point"])
        except:
            point = 0
        if point == 0:
            res.status_code = 200
            return res

        if jobTitle:
            userAccountChainyenList = UserAccountChainYenInfo.objects.filter(jobTitle=jobTitle)
            userAccountChainyenList.update(point=F('point') + point)
        else:
            res.status_code = 404
            return res

        res.status_code = 200

        modifier = request.user.user
        for userAccountChainyen in userAccountChainyenList:
            userAccount = userAccountChainyen.UserAccountInfo
            resultPoint = userAccountChainyen.point
            pHistory = pointHistory(UserAccountInfo = userAccount, modifier = modifier,
                                    recordDate = datetime.datetime.now(), reason = '管理者加點',
                                    addPoint = "+" + str(point), reducePoint = "", transferPoint = "",
                                    resultPoint = resultPoint)
            pHistory.save()

    except:
        res.status_code = 503

    return res

@permission_required('userlogin.seeManagerPointPage', login_url='/accounts/userlogin/')
def addPointByExcel(request):
    res = HttpResponse()
    try:
        excelFile = ""
        if "excelFile" in request.FILES:
            excelFile = request.FILES['excelFile']
        
        name = excelFile.name
        wb = load_workbook(excelFile)
        sheet = wb.active
        resContent = []
        modifier = request.user.user
        for i in range(2,sheet.max_row+1):
            if not sheet['B' + str(i)].value:
                break
            resTmp = {}
            resTmp["user"] = str(sheet["A" + str(i)].value)
            amwayNumber = str(sheet['B' + str(i)].value)
            id4 = str(sheet['C' + str(i)].value)
            username = amwayNumber + id4
            resTmp["amwayNumber"] = amwayNumber
            point = int(sheet['D' + str(i)].value)
            try:
                userAccountInfo = UserAccountInfo.objects.get(username = username)
                userAccountChainyen = UserAccountChainYenInfo.objects.filter(UserAccountInfo=userAccountInfo)
                userAccountChainyen.update(point=F('point') + point)
                resultPoint = userAccountChainyen.first().point
                pHistory = pointHistory(UserAccountInfo = userAccountInfo, modifier = modifier,
                                        recordDate = datetime.datetime.now(), reason = '管理者加點',
                                        addPoint = "+" + str(point), reducePoint = "", transferPoint = "",
                                        resultPoint = resultPoint)
                pHistory.save()
                resTmp["resultPoint"] = resultPoint
                resTmp["point"] = point
            except:
                resTmp["resultPoint"] = "加點失敗"
                resTmp["point"] = "加點失敗"
            resContent.append(resTmp)
        res.status_code = 200
        res.content = json.dumps(resContent)
    except:
        res.status_code = 503
    return res

@permission_required('userlogin.seeManagerPointPage', login_url='/accounts/userlogin/')
def reducePoint(request):
    res = HttpResponse()
    try:
        userAccountInfo = UserAccountInfo.objects.get(username = request.POST['username'])
        accountChainyen = UserAccountChainYenInfo.objects.get(UserAccountInfo = userAccountInfo)
        point = accountChainyen.point -1
        accountChainyen.point = F('point') - 1
        accountChainyen.save()

        res.status_code = 200
        res.content = point

        modifier = request.user.user
        pHistory = pointHistory(UserAccountInfo = userAccountInfo, modifier = modifier,
                                recordDate = datetime.datetime.now(), reason = '管理者扣點',
                                addPoint = "", reducePoint = "-1", transferPoint = "",
                                resultPoint = point)
        pHistory.save()
    except:
        res.status_code = 503

    return res

@permission_required('userlogin.seeManagerPointPage', login_url='/accounts/userlogin/')
def getPointHistory(request):
    res = HttpResponse()
    try:
        if request.method == "POST":
            userAccountInfo = UserAccountInfo.objects.get(username = request.POST['username'])
        elif request.method == "GET":
            userAccountInfo = UserAccountInfo.objects.get(username = request.user.username)

        pHistory = pointHistory.objects.filter(UserAccountInfo = userAccountInfo).order_by('-recordDate')
        res.status_code = 200
        res.content =  serializers.serialize("json", pHistory)
    except:
        res.status_code = 503

    return res

@permission_required('userlogin.seeManagerPointPage', login_url='/accounts/userlogin/')
def allUserAccount(request):
    res = HttpResponse()
    try:
        allData = []
        allUserAccountInfo = UserAccountInfo.objects.all()
        for user in allUserAccountInfo:
            userAmwayAccountInfo = UserAccountAmwayInfo.objects.get(UserAccountInfo=user)
            chainyenAccount = UserAccountChainYenInfo.objects.get(UserAccountInfo=user)
            temp = {
                "username": user.username,
                "user": user.user,
                "amwayNumber": userAmwayAccountInfo.amwayNumber,
                "jobTitle":chainyenAccount.jobTitle.jobTitle,
                "amwayAward": userAmwayAccountInfo.amwayAward.amwayAward,
                "point": chainyenAccount.point
            }
            allData.append(temp)
        res.status_code = 200
        res.content =  json.dumps(allData)
    except:
        res.status_code = 503

    return res


def getSelfPointHistory(request):
    res = HttpResponse()
    try:
        userAccountInfo = UserAccountInfo.objects.get(username = request.user.username)
        pHistory = pointHistory.objects.filter(UserAccountInfo = userAccountInfo).order_by('-recordDate')
        res.status_code = 200
        res.content =  serializers.serialize("json", pHistory)
    except:
        res.status_code = 503

    return res


def transferPoint(request):
    res = HttpResponse()
    try:

        fromUser = request.user.username
        toUser = request.POST['username']
        pointTransfer = int(request.POST['point'])

        #Do not hack me
        if pointTransfer <= 0:
            res.status_code = 404
            res.content = "do not try to hack me <3"
            return res

        #TODO restrict the transfer outside self Team when someone hack this API
        fromUserAccount = UserAccountInfo.objects.get(username = fromUser)
        fromUserAmwayAccount = UserAccountAmwayInfo.objects.get(UserAccountInfo = fromUserAccount)
        fromUserAmwayNumber = fromUserAmwayAccount.amwayNumber
        toUserAccount = UserAccountInfo.objects.get(username = toUser)
        toUserAmwayAccountInfo = UserAccountAmwayInfo.objects.get(UserAccountInfo=toUserAccount)
        toUserDDInfo = toUserAmwayAccountInfo.amwayDD

        #Do not hack me
        if not (toUserDDInfo.amwayNumber  == fromUserAmwayNumber or 
                int(toUserDDInfo.amwayDiamond) == fromUserAmwayNumber):
            res.status_code = 404
            res.content = "do not try to hack me <3"
            return res
        
        #扣點
        userAccountInfo = UserAccountInfo.objects.get(username = fromUser)
        accountChainyen = UserAccountChainYenInfo.objects.get(UserAccountInfo = userAccountInfo)
        fromResultPoint = accountChainyen.point - pointTransfer
        if fromResultPoint < 0:
            res.status_code = 404
            res.content = "point not enough"
            return res
        accountChainyen.point = F('point') - pointTransfer
        accountChainyen.save()

        modifier = request.user.user
        pHistory = pointHistory(UserAccountInfo = userAccountInfo, modifier = modifier,
                                recordDate = datetime.datetime.now(), reason = '轉讓點數',
                                addPoint = "", reducePoint = "", transferPoint = "-" + str(pointTransfer),
                                resultPoint = fromResultPoint)
        pHistory.save()

        #加點
        userAccountInfo = UserAccountInfo.objects.get(username = toUser)
        accountChainyen = UserAccountChainYenInfo.objects.get(UserAccountInfo = userAccountInfo)
        toResultPoint = accountChainyen.point + int(pointTransfer)
        accountChainyen.point = F('point') + int(pointTransfer)
        accountChainyen.save()

        res.status_code = 200
        res.content = json.dumps({
            "fromUser": fromUser,
            "toUser": toUser,
            "fromResultPoint": fromResultPoint,
            "toResultPoint": toResultPoint
        })

        modifier = request.user.user
        pHistory = pointHistory(UserAccountInfo = userAccountInfo, modifier = modifier,
                                recordDate = datetime.datetime.now(), reason = '轉讓點數',
                                addPoint = "", reducePoint = "", transferPoint = "+" + str(pointTransfer),
                                resultPoint = toResultPoint)
        pHistory.save()
    except:
        res.status_code = 503

    return res


def getPersonalTeam(request):
    res = HttpResponse()

    username = request.user.username
    userAccountInfo = UserAccountInfo.objects.get(username = username)

    try:
        #轉讓: 白金小組 & 鑽石小組
        myTeam = []
        userAccountAmway = UserAccountAmwayInfo.objects.get(UserAccountInfo = userAccountInfo)
        awardRank = userAccountAmway.amwayAward.rank
        amNumber  = userAccountAmway.amwayNumber

        #白金
        if awardRank >= 15 and awardRank < 60:
            myDDInfo = registerDDandDimInfo.objects.get(amwayNumber = amNumber)
            accountIDList = UserAccountAmwayInfo.objects.filter(amwayDD = myDDInfo).values('UserAccountInfo')
            myTeam = UserAccountInfo.objects.filter(id__in=accountIDList).exclude(username = request.user.username)
        #鑽石
        elif awardRank >= 60:
            myDDInfo = registerDDandDimInfo.objects.get(amwayNumber = amNumber)
            accountIDList = UserAccountAmwayInfo.objects.filter(amwayDD = myDDInfo).values('UserAccountInfo')
            myTeam = UserAccountInfo.objects.filter(id__in=accountIDList).exclude(username = request.user.username)
            myTeamDD = registerDDandDimInfo.objects.filter(amwayDiamond = str(amNumber))
            myTeam = list(myTeam)
            for myDD in myTeamDD:
                accountIDList = UserAccountAmwayInfo.objects.filter(amwayDD = myDD).values('UserAccountInfo')
                if accountIDList:
                    ddTeam = UserAccountInfo.objects.filter(id__in = accountIDList).exclude(username = username)
                    for item in ddTeam:
                        if item not in myTeam:
                            myTeam.append(item)
        else:
            myTeam = []
    except:
        # 此直銷權在資料庫沒有註冊為 DD
        myTeam = []
    
    finalTeam = []
    for userAccount in myTeam:
        amwayAccount = UserAccountAmwayInfo.objects.get(UserAccountInfo=userAccount)
        chainyenAccount = UserAccountChainYenInfo.objects.get(UserAccountInfo=userAccount)
        temp = {
            "username": userAccount.username,
            "user": userAccount.user,
            "amwayNumber": amwayAccount.amwayNumber,
            "point": chainyenAccount.point
        }
        finalTeam.append(temp)
    res.status_code = 200
    res.content = json.dumps(finalTeam)
    return res
