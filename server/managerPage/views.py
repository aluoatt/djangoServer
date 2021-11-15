import json

from django.http import HttpResponse
from django.contrib.auth.models import Permission
import traceback
from django.core import serializers


# Create your views here.
from . import models
from django.shortcuts import render
from django.db.models import Q

from userlogin.models import UserAccountInfo, UserAccountChainYenInfo, chainYenJobTitleInfo
from userlogin.models import chainYenClassInfo, registerDDandDimInfo, amwayAwardInfo, ConfirmString,UserAccountAmwayInfo
from django.contrib.auth.decorators import permission_required
from userlogin.models import TempUserAccountInfo, AccountModifyHistory, TempUserAccountChainYenInfo, TempUserAccountAmwayInfo
from NutriliteSearchPage.models import DBClassInfo, fileDataInfo, mainClassInfo, secClassInfo, articleModifyHistory, articleReport
from NutriliteSearchPage.models import fileDataKeywords, personalFileData, personalWatchFileLog
from openpyxl import load_workbook

# Create your views here.
import hashlib
import datetime
from django.conf import settings
def hash_code(s, salt='mysite'):
    h = hashlib.sha256()
    s += salt
    h.update(s.encode())
    return h.hexdigest()

def make_confirm_string(user):
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    code = hash_code(user, now)
    ConfirmStringObj = ConfirmString.objects.filter(user_name=user)
    if ConfirmStringObj.count() > 0 :
        ConfirmStringObj.first().delete()

    ConfirmString.objects.create(code=code, user_name=user)
    return code

def send_email(email, code):

    from django.core.mail import EmailMultiAlternatives

    subject = '來自群雁資訊檢索系統的確認信'

    text_content = '''
                    *此信件為系統發出信件，請勿直接回覆，感謝您的配合，謝謝!*
                    感謝您註冊自群雁資訊檢索系統！
                    如果你看到這則消息，說明你的信箱不提供HTML連接功能，請洽會長或上手白金！'''

    html_content = '''
                    <p>*此信件為系統發出信件，請勿直接回覆，感謝您的配合，謝謝!*<p>
                    <p>感謝您註冊自群雁資訊檢索系統</p>                  
                    <p>請點我認證註冊！</p>
                    <p><a href="{}/managerPages/userAccountConfirm?code={}" target=blank>確認連結</a></p>
                    <p>此連結的有效期為{}天！</p>
                    '''.format(settings.MYIP, code, settings.CONFIRM_DAYS)

    msg = EmailMultiAlternatives(subject, text_content, settings.EMAIL_HOST_USER, [email])
    msg.attach_alternative(html_content, "text/html")
    msg.send()

def userAccountConfirm(request):
    code = request.GET.get('code', None)
    message = ''
    try:
        confirm = ConfirmString.objects.get(code=code)
    except:
        message = '無效的確認請求!'
        return render(request, 'managerPages/userAccountConfirm.html', locals())

    c_time = confirm.c_time
    now = datetime.datetime.now()
    if now > c_time + datetime.timedelta(settings.CONFIRM_DAYS):
        confirm.delete()
        message = '您的郵件已經過期，請洽會長或上手白金!'
        return render(request, 'managerPages/userAccountConfirm.html', locals())
    else:
        tr = TempUserAccountInfo.objects.get(username=confirm.user_name)
        classChairMan = tr.tempuseraccountchainyeninfo_set.first().jobTitle.jobTitle == "會長"

        r = UserAccountInfo(username=tr.username,
                            user=tr.user,
                            gender=tr.gender,
                            phone=tr.phone,
                            password=tr.password,
                            is_superuser=0,
                            is_staff=0,
                            is_active=1,
                            dataPermissionsLevel=1,
                            email=tr.email)

        r2 = UserAccountChainYenInfo(UserAccountInfo=r,
                                     jobTitle=chainYenJobTitleInfo.objects.get(id=tr.tempuseraccountchainyeninfo_set.first().jobTitle.id),
                                     classRoom=chainYenClassInfo.objects.get(id=tr.tempuseraccountchainyeninfo_set.first().classRoom.id),
                                     accountStatus="正常",
                                     freezeDate=None,
                                     point=10,
                                     EM=tr.tempuseraccountchainyeninfo_set.first().EM)

        r3 = UserAccountAmwayInfo(UserAccountInfo=r,
                                  amwayNumber=tr.tempuseraccountamwayinfo_set.first().amwayNumber,
                                  amwayAward=amwayAwardInfo.objects.get(id=tr.tempuseraccountamwayinfo_set.first().amwayAward.id),
                                  amwayDD=registerDDandDimInfo.objects.get(amwayNumber=tr.tempuseraccountamwayinfo_set.first().amwayDD.amwayNumber)
                                  )

        r.save()
        r2.save()
        r3.save()

        if classChairMan:
            userAcconut = UserAccountInfo.objects.get(username=tr.username)
            perm = Permission.objects.get(codename=settings.CLASS_CHARIMAN_MANAGER_DICT[tr.tempuseraccountchainyeninfo_set.first().classRoom.ClassRoomName])
            userAcconut.user_permissions.add(perm)
            perm = Permission.objects.get(codename="can_see_register")
            userAcconut.user_permissions.add(perm)
            perm = Permission.objects.get(codename="seeManagerMenuButton")
            userAcconut.user_permissions.add(perm)
            perm = Permission.objects.get(codename="seeManagerAccountManagerPage")
            userAcconut.user_permissions.add(perm)
            perm = Permission.objects.get(codename="seeManagerAuditAccountPage")
            userAcconut.user_permissions.add(perm)
            perm = Permission.objects.get(codename="can_Change_JobTitle")
            userAcconut.user_permissions.add(perm)
        tr.delete()
        confirm.delete()
        message = '恭喜您註冊成功!五秒後將自動跳轉到登入頁面。'
        return render(request, 'managerPages/userAccountConfirm.html', locals())

@permission_required('userlogin.seeManagerAccountManagerPage', login_url='/accounts/userlogin/')
def managerAccountManagerPage(request):
    tag = "ManagerAccountManagerPage"

    # 職務表
    jobTitles = chainYenJobTitleInfo.objects.all()
    # 獎銜表
    amwayAwards = amwayAwardInfo.objects.all().order_by('rank')
    # 教室表
    chainYenClasses = chainYenClassInfo.objects.all().order_by('rank')
    # 白金表
    registerDDs = registerDDandDimInfo.objects.filter(amwayAward__rank__gte=15).order_by('amwayNumber')
    # 鑽石表
    registerDims = registerDDandDimInfo.objects.filter(amwayAward__rank__gte=60).order_by('amwayNumber')



    q1 = Q()
    q1.connector = 'OR'
    q1.children.append(("classRoom__ClassRoomName", "無"))

    if request.user.has_perm('userlogin.CYPManager'):
        q1.children.append(("classRoom__ClassRoomName", "台北"))

    if request.user.has_perm('userlogin.CYLManager'):
        q1.children.append(("classRoom__ClassRoomName", "中壢"))

    if request.user.has_perm('userlogin.CYSManager'):
        q1.children.append(("classRoom__ClassRoomName", "新竹"))

    if request.user.has_perm('userlogin.CYZManager'):
        q1.children.append(("classRoom__ClassRoomName", "台中"))

    if request.user.has_perm('userlogin.CYJManager'):
        q1.children.append(("classRoom__ClassRoomName", "嘉義"))

    if request.user.has_perm('userlogin.CYN2Manager'):
        q1.children.append(("classRoom__ClassRoomName", "永康245"))

    if request.user.has_perm('userlogin.CYN1Manager'):
        q1.children.append(("classRoom__ClassRoomName", "永康135"))

    if request.user.has_perm('userlogin.CYMManager'):
        q1.children.append(("classRoom__ClassRoomName", "良美"))

    if request.user.has_perm('userlogin.CYKManager'):
        q1.children.append(("classRoom__ClassRoomName", "高雄"))

    if request.user.has_perm('userlogin.CYDManager'):
        q1.children.append(("classRoom__ClassRoomName", "屏東"))

    if request.user.has_perm('userlogin.CYWManager'):
        q1.children.append(("classRoom__ClassRoomName", "花蓮"))

    if request.user.has_perm('userlogin.CYTManager'):
        q1.children.append(("classRoom__ClassRoomName", "台東"))

    if request.user.has_perm('userlogin.CYHManager'):
        q1.children.append(("classRoom__ClassRoomName", "澎湖"))

    q2 = Q()
    q2.connector = 'OR'
    q2.children.append(("id", 0))
    for UserAccountChainYen in UserAccountChainYenInfo.objects.filter(q1):
        q2.children.append(("id", UserAccountChainYen.UserAccountInfo.id))
    # UserAccountChainYen = UserAccountChainYenInfo.objects.filter(q1)
    UserAccount = UserAccountInfo.objects.get(username=request.user)
    if registerDDandDimInfo.objects.filter(amwayNumber = UserAccount.useraccountamwayinfo_set.first().amwayNumber).count() > 0:
        for UserAccountAmway in UserAccountAmwayInfo.objects.filter(amwayDD=registerDDandDimInfo.objects.get(amwayNumber = UserAccount.useraccountamwayinfo_set.first().amwayNumber).id):
            q2.children.append(("id", UserAccountAmway.UserAccountInfo.id))
    searchUserAccountInfo = UserAccountInfo.objects.filter(q2).order_by('username')

    return render(request, 'managerPages/managerAccountManagerPage.html', locals())

#由管理者修改個人資料
@permission_required('userlogin.seeManagerAccountManagerPage', login_url='/accounts/userlogin/')
def managerAccountModify(request):
    tag = "ManagerAuditAccountPage"

    # 職務表
    jobTitles = chainYenJobTitleInfo.objects.all()
    # 獎銜表
    amwayAwards = amwayAwardInfo.objects.all().order_by('rank')
    # 教室表
    chainYenClasses = chainYenClassInfo.objects.all().order_by('rank')
    # 白金表
    registerDDs = registerDDandDimInfo.objects.filter(amwayAward__rank__gte=15).order_by('amwayNumber')
    # 鑽石表
    registerDims = registerDDandDimInfo.objects.filter(amwayAward__rank__gte=60).order_by('amwayNumber')

    TempUserAccountChainYen = TempUserAccountChainYenInfo.objects.all()
    # print(TempUserAccount)
    # kwargs = {}
    # 台北
    if not request.user.has_perm('userlogin.CYPManager'):
        TempUserAccountChainYen = TempUserAccountChainYen.exclude(classRoom__ClassRoomName="台北")

    if not request.user.has_perm('userlogin.CYLManager'):
        TempUserAccountChainYen = TempUserAccountChainYen.exclude(classRoom__ClassRoomName="中壢")

    if not request.user.has_perm('userlogin.CYSManager'):
        TempUserAccountChainYen = TempUserAccountChainYen.exclude(classRoom__ClassRoomName="新竹")

    if request.user.has_perm('userlogin.CYZManager'):
        TempUserAccountChainYen = TempUserAccountChainYen.exclude(classRoom__ClassRoomName="台中")

    if request.user.has_perm('userlogin.CYJManager'):
        TempUserAccountChainYen = TempUserAccountChainYen.exclude(classRoom__ClassRoomName="嘉義")
    if request.user.has_perm('userlogin.CYN2Manager'):
        TempUserAccountChainYen = TempUserAccountChainYen.exclude(classRoom__ClassRoomName="永康245")
    if request.user.has_perm('userlogin.CYN1Manager'):
        TempUserAccountChainYen = TempUserAccountChainYen.exclude(classRoom__ClassRoomName="永康135")
    if request.user.has_perm('userlogin.CYMManager'):
        TempUserAccountChainYen = TempUserAccountChainYen.exclude(classRoom__ClassRoomName="良美")
    if request.user.has_perm('userlogin.CYKManager'):
        TempUserAccountChainYen = TempUserAccountChainYen.exclude(classRoom__ClassRoomName="高雄")
    if request.user.has_perm('userlogin.CYDManager'):
        TempUserAccountChainYen = TempUserAccountChainYen.exclude(classRoom__ClassRoomName="屏東")
    if request.user.has_perm('userlogin.CYWManager'):
        TempUserAccountChainYen = TempUserAccountChainYen.exclude(classRoom__ClassRoomName="花蓮")
    if request.user.has_perm('userlogin.CYTManager'):
        TempUserAccountChainYen = TempUserAccountChainYen.exclude(classRoom__ClassRoomName="台東")
    if request.user.has_perm('userlogin.CYHManager'):
        TempUserAccountChainYen = TempUserAccountChainYen.exclude(classRoom__ClassRoomName="澎湖")
    return render(request, 'managerPages/managerAuditAccountManagerPage.html', locals())


@permission_required('userlogin.seeManagerAccountManagerPage', login_url='/accounts/userlogin/')
def modalAccountModifyPOST(request):
    # AccountModifyHistory
    try:
        managerName = UserAccountInfo.objects.get(username=request.user).user
        userid = json.loads(request.body.decode('utf-8'))["userid"]

        user = json.loads(request.body.decode('utf-8'))["modal_user"]
        phone = json.loads(request.body.decode('utf-8'))["phone"]
        # amwayNumber = json.loads(request.body.decode('utf-8'))["modal_amwayNumber"]
        gender = json.loads(request.body.decode('utf-8'))["gender"]
        if request.user.has_perm('userlogin.can_Change_DataPermission'):
            dataPermissionsLevel = json.loads(request.body.decode('utf-8'))["modal_dataPermissionsLevel"]

        if request.user.has_perm('userlogin.can_Change_JobTitle'):
            chainYenJobTitle = json.loads(request.body.decode('utf-8'))["chainYenJobTitle"]


        if request.user.has_perm('userlogin.can_Change_class'):
            classRoom = json.loads(request.body.decode('utf-8'))["classRoom"]

        if request.user.has_perm('userlogin.can_freeze_account'):
            accountStatus = json.loads(request.body.decode('utf-8'))["modal_accountStatus"]
            if accountStatus == "正常":
                accountStatus = True
            else:
                accountStatus = False

        amwayAward = json.loads(request.body.decode('utf-8'))["modal_amwayAward"]
        amwayDD = json.loads(request.body.decode('utf-8'))["amwayDD"]

        r = UserAccountInfo.objects.get(id=int(userid))
        if r.user != user:

            AccountModifyHistory(UserAccountInfo=r,
                                 modifier = managerName,
                                 recordDate = str(datetime.datetime.now()),
                                 modifyFielddName = "姓名",
                                 originFieldData = r.user,
                                 RevisedData = user).save()
            r.user = user

        if r.gender != gender:
            AccountModifyHistory(UserAccountInfo=r,
                                 modifier=managerName,
                                 recordDate=str(datetime.datetime.now()),
                                 modifyFielddName="性別",
                                 originFieldData=r.gender,
                                 RevisedData=gender).save()
            r.gender = gender

        if r.phone != phone:
            AccountModifyHistory(UserAccountInfo=r,
                                 modifier=managerName,
                                 recordDate=str(datetime.datetime.now()),
                                 modifyFielddName="電話",
                                 originFieldData=r.phone,
                                 RevisedData=phone).save()
            r.phone = phone

        if request.user.has_perm('userlogin.can_Change_DataPermission'):

            if r.dataPermissionsLevel != int(dataPermissionsLevel):
                AccountModifyHistory(UserAccountInfo=r,
                                     modifier=managerName,
                                     recordDate=str(datetime.datetime.now()),
                                     modifyFielddName="資料權限等級",
                                     originFieldData=r.dataPermissionsLevel,
                                     RevisedData=int(dataPermissionsLevel)).save()

                r.dataPermissionsLevel = int(dataPermissionsLevel)

        if request.user.has_perm('userlogin.can_freeze_account'):

            if r.is_active != accountStatus:
                if accountStatus:

                    AccountModifyHistory(UserAccountInfo=r,
                                         modifier=managerName,
                                         recordDate=str(datetime.datetime.now()),
                                         modifyFielddName="資料權限等級",
                                         originFieldData="凍結",
                                         RevisedData="正常").save()
                else:
                    AccountModifyHistory(UserAccountInfo=r,
                                        modifier=managerName,
                                        recordDate=str(datetime.datetime.now()),
                                        modifyFielddName="帳號狀態",
                                        originFieldData="正常",
                                        RevisedData="凍結").save()

                r.is_active = accountStatus


        r2 = UserAccountChainYenInfo.objects.get(UserAccountInfo = int(userid))
        if request.user.has_perm('userlogin.can_Change_JobTitle'):
            odata = chainYenJobTitleInfo.objects.get(jobTitle=chainYenJobTitle)
            if r2.jobTitle.jobTitle != odata.jobTitle:
                AccountModifyHistory(UserAccountInfo=r,
                                     modifier=managerName,
                                     recordDate=str(datetime.datetime.now()),
                                     modifyFielddName="職務",
                                     originFieldData=r2.jobTitle.jobTitle,
                                     RevisedData=odata.jobTitle).save()

                r2.jobTitle = odata

            if odata.jobTitle == "會長":

                userAcconut = r
                perm = Permission.objects.get(codename=settings.CLASS_CHARIMAN_MANAGER_DICT[
                    userAcconut.useraccountchainyeninfo_set.first().classRoom.ClassRoomName])
                userAcconut.user_permissions.add(perm)
                perm = Permission.objects.get(codename="can_see_register")
                userAcconut.user_permissions.add(perm)
                perm = Permission.objects.get(codename="seeManagerMenuButton")
                userAcconut.user_permissions.add(perm)
                perm = Permission.objects.get(codename="seeManagerAccountManagerPage")
                userAcconut.user_permissions.add(perm)
                perm = Permission.objects.get(codename="seeManagerAuditAccountPage")
                userAcconut.user_permissions.add(perm)
                perm = Permission.objects.get(codename="can_Change_JobTitle")
                userAcconut.user_permissions.add(perm)
            else:
                userAcconut = r
                perm = Permission.objects.get(codename="can_see_register")
                userAcconut.user_permissions.remove(perm)
                perm = Permission.objects.get(codename="seeManagerMenuButton")
                userAcconut.user_permissions.remove(perm)
                perm = Permission.objects.get(codename="seeManagerAccountManagerPage")
                userAcconut.user_permissions.remove(perm)
                perm = Permission.objects.get(codename="seeManagerAuditAccountPage")
                userAcconut.user_permissions.remove(perm)
                perm = Permission.objects.get(codename="can_Change_JobTitle")
                userAcconut.user_permissions.remove(perm)



        if request.user.has_perm('userlogin.can_Change_class'):


            odata = chainYenClassInfo.objects.get(ClassRoomName = classRoom)
            if r2.classRoom.ClassRoomName != odata.ClassRoomName:
                AccountModifyHistory(UserAccountInfo=r,
                                     modifier=managerName,
                                     recordDate=str(datetime.datetime.now()),
                                     modifyFielddName="教室",
                                     originFieldData=r2.classRoom.ClassRoomName,
                                     RevisedData=odata.ClassRoomName).save()

                r2.classRoom = odata

        r3 = UserAccountAmwayInfo.objects.get(UserAccountInfo=int(userid))

        odata = amwayAwardInfo.objects.get(amwayAward = amwayAward)
        if r3.amwayAward.amwayAward != odata.amwayAward:
            AccountModifyHistory(UserAccountInfo=r,
                                 modifier=managerName,
                                 recordDate=str(datetime.datetime.now()),
                                 modifyFielddName="教室",
                                 originFieldData=r3.amwayAward.amwayAward,
                                 RevisedData=odata.amwayAward).save()

            r3.amwayAward = amwayAwardInfo.objects.get(amwayAward = amwayAward)

        odata = registerDDandDimInfo.objects.get(amwayNumber=int(amwayDD))
        if r3.amwayDD.amwayNumber != odata.amwayNumber:
            AccountModifyHistory(UserAccountInfo=r,
                                 modifier=managerName,
                                 recordDate=str(datetime.datetime.now()),
                                 modifyFielddName="教室",
                                 originFieldData=r3.amwayDD.amwayNumber,
                                 RevisedData=odata.amwayNumber).save()

            r3.amwayDD = odata


        r.save()
        r2.save()
        r3.save()
    except:
        print(traceback.print_exc())
        response_data={}
        response_data["status"] = False
        return HttpResponse(json.dumps(response_data), content_type="application/json")

    response_data = {}
    response_data["status"] = True
    return HttpResponse(json.dumps(response_data), content_type="application/json")

@permission_required('userlogin.seeManagerAccountManagerPage', login_url='/accounts/userlogin/')
def changeStatusByExcel(request):
    res = HttpResponse()
    try:
        excelFile = ""
        if "excelFile" in request.FILES:
            excelFile = request.FILES['excelFile']
        
        is_active = ""
        if "is_active" in request.POST:
            is_active = int(request.POST['is_active'])
        
        point = 0
        if "point" in request.POST:
            point = int(request.POST["point"])
        
        if excelFile == "" or is_active == "" or \
            (is_active == True and is_active == False):
            res.content = "No info provide"
            res.status_code = 400
            return res

        if not request.user.has_perm('userlogin.can_freeze_account'):
            res.content = "Permission not enough"
            res.status_code = 400
            return res

        wb = load_workbook(excelFile)
        sheet = wb.active
        resContent = []
        modifier = request.user.user
        for i in range(2,sheet.max_row+1):
            if not sheet['B' + str(i)].value:
                break
            resTmp = {}
            resTmp["user"] = str(sheet["A" + str(i)].value)
            amwayNumber = str(sheet['B' + str(i)].value)
            id4 = str(sheet['C' + str(i)].value)
            username = amwayNumber + id4
            resTmp["amwayNumber"] = amwayNumber
            r = UserAccountInfo.objects.get(username=int(username))
            if r.is_active != is_active:
                if is_active:
                    AccountModifyHistory(UserAccountInfo=r,
                                            modifier=modifier,
                                            recordDate=str(datetime.datetime.now()),
                                            modifyFielddName="帳號狀態",
                                            originFieldData="凍結",
                                            RevisedData="正常").save()
                    rChainyen = UserAccountChainYenInfo.objects.get(UserAccountInfo=r)
                    rChainyen.point = point
                    rChainyen.save()
                    resTmp['point'] = point
                else:
                    AccountModifyHistory(UserAccountInfo=r,
                                        modifier=modifier,
                                        recordDate=str(datetime.datetime.now()),
                                        modifyFielddName="帳號狀態",
                                        originFieldData="正常",
                                        RevisedData="凍結").save()
                    rChainyen = UserAccountChainYenInfo.objects.get(UserAccountInfo=r)
                    rChainyen.point = 0
                    rChainyen.save()
                    resTmp['point'] = 0

                r.is_active = is_active
                r.save()
                if is_active:
                    resTmp['status'] = "啟用"
                else:
                    resTmp['status'] = "凍結"
            else:
                if is_active:
                    resTmp['status'] = "原本就為啟用"
                else:
                    resTmp['status'] = "原本就為凍結"
                resTmp["point"] = "不變"

            resContent.append(resTmp)
        res.status_code = 200
        res.content = json.dumps(resContent)
    except:
        res.status_code = 503
    return res


@permission_required('userlogin.seeManagerAuditAccountPage', login_url='/accounts/userlogin/')
def managerAuditAccountPage(request):
    tag = "ManagerAuditAccountPage"

    # 職務表
    jobTitles = chainYenJobTitleInfo.objects.all()
    # 獎銜表
    amwayAwards = amwayAwardInfo.objects.all().order_by('rank')
    # 教室表
    chainYenClasses = chainYenClassInfo.objects.all().order_by('rank')
    # 白金表
    registerDDs = registerDDandDimInfo.objects.filter(amwayAward__rank__gte=15).order_by('amwayNumber')
    # 鑽石表
    registerDims = registerDDandDimInfo.objects.filter(amwayAward__rank__gte=60).order_by('amwayNumber')

    return render(request, 'managerPages/managerAuditAccountManagerPage.html', locals())

@permission_required('userlogin.seeManagerAuditAccountPage', login_url='/accounts/userlogin/')
def removeAuditAccount(request):
    response_data = {}

    try:
        tempAccountId = int(json.loads(request.body.decode('utf-8'))["id"])
        userAccount = TempUserAccountInfo.objects.filter(id=tempAccountId)
        confirm = ConfirmString.objects.filter(user_name=userAccount.first().username)
        if confirm.count()>0:
            confirm.delete()

        if userAccount.count()>0:
            userAccount.first().delete()

        response_data["status"] = True
    except:
        response_data["status"] = False
        return HttpResponse(json.dumps(response_data), content_type="application/json")


    return HttpResponse(json.dumps(response_data), content_type="application/json")

@permission_required('userlogin.seeManagerAuditAccountPage', login_url='/accounts/userlogin/')
def AcceptAuditAccount(request):
    response_data = {}

    try:
        tempAccountId = int(json.loads(request.body.decode('utf-8'))["id"])
        TempUserAccount = TempUserAccountInfo.objects.get(id=tempAccountId)
        code = make_confirm_string(TempUserAccount.username)
        send_email(TempUserAccount.email, code)

        TempUserAccount.auditStatus = "已寄信"
        TempUserAccount.save()
        response_data["status"] = True

    except:
        response_data["status"] = False
        return HttpResponse(json.dumps(response_data), content_type="application/json")

    return HttpResponse(json.dumps(response_data), content_type="application/json")

@permission_required('userlogin.seeManagerPointPage', login_url='/accounts/userlogin/')
def managerPointManagerPage(request):
    tag = "ManagerPointManagerPage"

    # 職務表
    jobTitles = chainYenJobTitleInfo.objects.all()
    jobTitleList = []
    for jobTitle in jobTitles:
        jobTitleList.append(jobTitle.jobTitle)
    jobTitleList = json.dumps(jobTitleList)
    # 獎銜表
    amwayAwards = amwayAwardInfo.objects.all().order_by('rank')
    amwayAwardList = []
    for amAward in amwayAwards:
        amwayAwardList.append(amAward.amwayAward)
    amwayAwardList = json.dumps(amwayAwardList)

    searchUserAccountInfo = UserAccountInfo.objects.all()
    return render(request, 'managerPages/managerPointManagerPage.html', locals())


@permission_required('userlogin.seeManagerAccountManagerPage', login_url='/accounts/userlogin/')
def getAccountModifyHistory(request):
    res = HttpResponse()
    # try:
    userAccountInfo = UserAccountInfo.objects.get(id=request.POST['username'])
    pHistory = AccountModifyHistory.objects.filter(UserAccountInfo=userAccountInfo).order_by('-recordDate')
    res.status_code = 200
    res.content = serializers.serialize("json", pHistory)
    # except:
    #     res.status_code = 503

    return res

@permission_required('userlogin.seeManagerArticlePage', login_url='/accounts/userlogin/')
def managerArticleManagerPage(request):
    tag = "ManagerArticleManagerPage"
    
    # 主類別
    mainClass = mainClassInfo.objects.all()
    mainClassList = []
    for mClass in mainClass:
        mainClassList.append(mClass.mainClassName)
    mainClassList = json.dumps(mainClassList)

    # 副類別
    secClass = secClassInfo.objects.all()
    secClassList = []
    for sClass in secClass:
        secClassList.append(sClass.secClassName)
    secClassList = json.dumps(secClassList)

    return render(request, 'managerPages/managerArticleManagerPage.html', locals())

@permission_required('userlogin.seeManagerArticleReportPage', login_url='/accounts/userlogin/')
def managerArticleReportManagerPage(request):
    tag = "ManagerArticleReportManagerPage"
    
    # 主類別
    mainClass = mainClassInfo.objects.all()
    mainClassList = []
    for mClass in mainClass:
        mainClassList.append(mClass.mainClassName)
    mainClassList = json.dumps(mainClassList)

    # 副類別
    secClass = secClassInfo.objects.all()
    secClassList = []
    for sClass in secClass:
        secClassList.append(sClass.secClassName)
    secClassList = json.dumps(secClassList)

    return render(request, 'managerPages/managerArticleReportManagerPage.html', locals())

@permission_required('userlogin.seeManagerStatisticPage', login_url='/accounts/userlogin/')
def managerStatisticManagerPage(request):
    tag = "StatisticsManagerPage"
    return render(request, 'managerPages/managerStatisticManagerPage.html', locals())

@permission_required('userlogin.seeManagerAccountManagerPage', login_url='/accounts/userlogin/')
def allUserAccount(request):
    res = HttpResponse()
    try:
        allData = []
        allUserAccountInfo = UserAccountInfo.objects.all()

        q1 = Q()
        q1.connector = 'OR'
        q1.children.append(("classRoom__ClassRoomName", "無"))

        if request.user.has_perm('userlogin.CYPManager'):
            q1.children.append(("classRoom__ClassRoomName", "台北"))

        if request.user.has_perm('userlogin.CYLManager'):
            q1.children.append(("classRoom__ClassRoomName", "中壢"))

        if request.user.has_perm('userlogin.CYSManager'):
            q1.children.append(("classRoom__ClassRoomName", "新竹"))

        if request.user.has_perm('userlogin.CYZManager'):
            q1.children.append(("classRoom__ClassRoomName", "台中"))

        if request.user.has_perm('userlogin.CYJManager'):
            q1.children.append(("classRoom__ClassRoomName", "嘉義"))

        if request.user.has_perm('userlogin.CYN2Manager'):
            q1.children.append(("classRoom__ClassRoomName", "永康245"))

        if request.user.has_perm('userlogin.CYN1Manager'):
            q1.children.append(("classRoom__ClassRoomName", "永康135"))

        if request.user.has_perm('userlogin.CYMManager'):
            q1.children.append(("classRoom__ClassRoomName", "良美"))

        if request.user.has_perm('userlogin.CYKManager'):
            q1.children.append(("classRoom__ClassRoomName", "高雄"))

        if request.user.has_perm('userlogin.CYDManager'):
            q1.children.append(("classRoom__ClassRoomName", "屏東"))

        if request.user.has_perm('userlogin.CYWManager'):
            q1.children.append(("classRoom__ClassRoomName", "花蓮"))

        if request.user.has_perm('userlogin.CYTManager'):
            q1.children.append(("classRoom__ClassRoomName", "台東"))

        if request.user.has_perm('userlogin.CYHManager'):
            q1.children.append(("classRoom__ClassRoomName", "澎湖"))

        q2 = Q()
        q2.connector = 'OR'
        q2.children.append(("id", 0))
        for UserAccountChainYen in UserAccountChainYenInfo.objects.filter(q1):
            q2.children.append(("id", UserAccountChainYen.UserAccountInfo.id))
        # UserAccountChainYen = UserAccountChainYenInfo.objects.filter(q1)
        UserAccount = UserAccountInfo.objects.get(username=request.user)
        if registerDDandDimInfo.objects.filter(
                amwayNumber=UserAccount.useraccountamwayinfo_set.first().amwayNumber).count() > 0:
            print(UserAccountAmwayInfo.objects.filter(amwayDD=UserAccount.useraccountamwayinfo_set.first().id))
            for UserAccountAmway in UserAccountAmwayInfo.objects.filter(amwayDD=registerDDandDimInfo.objects.get(
                    amwayNumber=UserAccount.useraccountamwayinfo_set.first().amwayNumber).id):
                q2.children.append(("id", UserAccountAmway.UserAccountInfo.id))
        searchUserAccountInfo = UserAccountInfo.objects.filter(q2).order_by('username')

        for user in searchUserAccountInfo:
            userAmwayAccountInfo = UserAccountAmwayInfo.objects.get(UserAccountInfo=user)
            chainyenAccount = UserAccountChainYenInfo.objects.get(UserAccountInfo=user)
            temp = {
                "id":user.id,
                "username": user.username,
                "user": user.user,
                "gender": user.gender,
                "phone": user.phone,
                "ClassRoomName": chainyenAccount.classRoom.ClassRoomName,
                "amwayNumber": userAmwayAccountInfo.amwayNumber,
                "jobTitle":chainyenAccount.jobTitle.jobTitle,
                "amwayDD": userAmwayAccountInfo.amwayDD.main,
                "dataPermissionsLevel":user.dataPermissionsLevel,
                "amwayDiamond": userAmwayAccountInfo.amwayDD.amwayDiamond,
                "amwayDD_number":userAmwayAccountInfo.amwayDD.amwayNumber,
                "amwayAward": userAmwayAccountInfo.amwayAward.amwayAward,
                "point": chainyenAccount.point,
                "accountStatus":user.is_active,
            }
            allData.append(temp)
        res.status_code = 200
        res.content =  json.dumps(allData)
    except:
        res.status_code = 503
        print(traceback.print_exc())
    return res
@permission_required('userlogin.seeManagerAuditAccountPage', login_url='/accounts/userlogin/')
def getTempUserAccount(request):
    res = HttpResponse()
    try:
        TempUserAccountChainYen = TempUserAccountChainYenInfo.objects.all()
        # print(TempUserAccount)
        # kwargs = {}
        # 台北
        if not request.user.has_perm('userlogin.CYPManager'):
            TempUserAccountChainYen = TempUserAccountChainYen.exclude(classRoom__ClassRoomName="台北")

        if not request.user.has_perm('userlogin.CYLManager'):
            TempUserAccountChainYen = TempUserAccountChainYen.exclude(classRoom__ClassRoomName="中壢")

        if not request.user.has_perm('userlogin.CYSManager'):
            TempUserAccountChainYen = TempUserAccountChainYen.exclude(classRoom__ClassRoomName="新竹")

        if not request.user.has_perm('userlogin.CYZManager'):
            TempUserAccountChainYen = TempUserAccountChainYen.exclude(classRoom__ClassRoomName="台中")

        if not request.user.has_perm('userlogin.CYJManager'):
            TempUserAccountChainYen = TempUserAccountChainYen.exclude(classRoom__ClassRoomName="嘉義")

        if not request.user.has_perm('userlogin.CYN2Manager'):
            TempUserAccountChainYen = TempUserAccountChainYen.exclude(classRoom__ClassRoomName="永康245")

        if not request.user.has_perm('userlogin.CYN1Manager'):
            TempUserAccountChainYen = TempUserAccountChainYen.exclude(classRoom__ClassRoomName="永康135")

        if not request.user.has_perm('userlogin.CYMManager'):
            TempUserAccountChainYen = TempUserAccountChainYen.exclude(classRoom__ClassRoomName="良美")

        if not request.user.has_perm('userlogin.CYKManager'):
            TempUserAccountChainYen = TempUserAccountChainYen.exclude(classRoom__ClassRoomName="高雄")

        if not request.user.has_perm('userlogin.CYDManager'):
            TempUserAccountChainYen = TempUserAccountChainYen.exclude(classRoom__ClassRoomName="屏東")

        if not request.user.has_perm('userlogin.CYWManager'):
            TempUserAccountChainYen = TempUserAccountChainYen.exclude(classRoom__ClassRoomName="花蓮")

        if not request.user.has_perm('userlogin.CYTManager'):
            TempUserAccountChainYen = TempUserAccountChainYen.exclude(classRoom__ClassRoomName="台東")

        if not request.user.has_perm('userlogin.CYHManager'):
            TempUserAccountChainYen = TempUserAccountChainYen.exclude(classRoom__ClassRoomName="澎湖")
        allData = []
        for tempUser in TempUserAccountChainYen:
            user = tempUser.UserAccountInfo
            userAmwayAccountInfo = TempUserAccountAmwayInfo.objects.get(UserAccountInfo=user)
            chainyenAccount = tempUser
            temp = {
                "id":user.id,
                "username": user.username,
                "user": user.user,
                "gender": user.gender,
                "phone": user.phone,
                "email": user.email,
                "auditStatus":user.auditStatus,
                "ClassRoomName": chainyenAccount.classRoom.ClassRoomName,
                "amwayNumber": userAmwayAccountInfo.amwayNumber,
                "jobTitle":chainyenAccount.jobTitle.jobTitle,
                "amwayDD": userAmwayAccountInfo.amwayDD.main,
                "amwayAward": userAmwayAccountInfo.amwayAward.amwayAward,
            }
            allData.append(temp)
        res.status_code = 200
        res.content =  json.dumps(allData)
    except:
        res.status_code = 503

    return res

#統計報表相關 API
@permission_required('userlogin.seeManagerStatisticPage', login_url='/accounts/userlogin/')
def getFileDataSummary(request):
    res = HttpResponse()
    fileDatas = fileDataInfo.objects.all()
    mainClassSummary = {}
    for data in fileDatas:
        mainClass = data.mainClass.mainClassName
        if mainClass in mainClassSummary:
            mainClassSummary[mainClass] = mainClassSummary[mainClass] + 1
        else:
            mainClassSummary[mainClass] = 1
    res.content = json.dumps(mainClassSummary)
    res.status_code = 200
    return res

@permission_required('userlogin.seeManagerStatisticPage', login_url='/accounts/userlogin/')
def getArticleOwnRank(request):
    res = HttpResponse()
    fileDatas = fileDataInfo.objects.filter(exchangeCount__gte=1)
    ownDataSummary = {}
    for data in fileDatas:
        fileID = data.id
        ownDataSummary[fileID] = {
            "fileID":      fileID,
            "title":       data.title,
            "mainClass":   data.mainClass.mainClassName,
            "costPoint":   data.point,
            "total":       data.exchangeCount, 
            "totalStars":  data.stars,
        }
    ownDataSummaryList = []
    for key in ownDataSummary:
        ownDataSummaryList.append(ownDataSummary[key])
    res.content = json.dumps(ownDataSummaryList)
    res.status_code = 200
    return res

@permission_required('userlogin.seeManagerStatisticPage', login_url='/accounts/userlogin/')
def getArticleWatchRank(request):
    res = HttpResponse()
    fileDatas = personalWatchFileLog.objects.all()
    mainClassSummary = {}
    for data in fileDatas:
        mainClass = data.mainClass.mainClassName
        if mainClass in mainClassSummary:
            mainClassSummary[mainClass] = mainClassSummary[mainClass] + 1
        else:
            mainClassSummary[mainClass] = 1
    res.content = json.dumps(mainClassSummary)
    res.status_code = 200
    return res

@permission_required('userlogin.seeManagerStatisticPage', login_url='/accounts/userlogin/')
def getPointOwnAll(request):
    res = HttpResponse()
    accountChainyen = UserAccountChainYenInfo.objects.all()
    result = []
    for account in accountChainyen:
        if account.accountStatus == "正常":
            accountAmway = UserAccountAmwayInfo.objects.get(UserAccountInfo=account.UserAccountInfo)
            amwayDiamond = registerDDandDimInfo.objects.get(amwayNumber = accountAmway.amwayDD.amwayDiamond)
            temp = {
                "user"  : account.UserAccountInfo.user,
                "amwayDiamond" : amwayDiamond.main,
                "classRoom" : account.classRoom.ClassRoomName,
                "point" : account.point,
            }
            result.append(temp)
    res.content = json.dumps(result)
    res.status_code = 200
    return res

@permission_required('userlogin.seeManagerArticlePage', login_url='/accounts/userlogin/')
def getFileDataInfo(request):
    res = HttpResponse()
    try:
        mainClassList = []
        if request.user.has_perm('userlogin.NutrilliteArticleManage'):
            mainClassList.append("營養")
        if request.user.has_perm('userlogin.ArtistryArticleManage'):
            mainClassList.append("美容")
        if request.user.has_perm('userlogin.TechArticleManage'):
            mainClassList.append("科技")
        if request.user.has_perm('userlogin.AmwayQueenArticleManage'):
            mainClassList.append("金鍋")
        if request.user.has_perm('userlogin.OtherArticleManage'):
            mainClassList.append("其他")
        if request.user.has_perm('userlogin.ChainyenArticleManage'):
            mainClassList.append("總部會議/活動")
        if request.user.has_perm('userlogin.SpeechArticleManage'):
            mainClassList.append("演講廳")

        fileDatas = fileDataInfo.objects.filter(mainClass__mainClassName__in=mainClassList)
        fileDataSummary = []
        for data in fileDatas:
            keywordResult = fileDataKeywords.objects.filter(fileDataInfoID = data)
            keywordList = []
            for item in keywordResult:
                keywordList.append(item.keyword)
            tmp = {
                "id"       : data.id,
                "title"    : data.title,
                "DBClassCode"  : data.DBClass.DBClassCode,
                "mainClass": data.mainClass.mainClassName,
                "secClass" : data.secClass.secClassName,
                "keyword"  : "#" + '#'.join(keywordList),
                "describe" : data.describe,
                "point"    : data.point,
                "visible"  : data.visible
            }
            fileDataSummary.append(tmp)
        res.content = json.dumps(fileDataSummary)
        res.status_code = 200
    except:
        res.status_code = 503
    return res

@permission_required('userlogin.seeManagerArticlePage', login_url='/accounts/userlogin/')
def getFileDataInfoByID(request, articleID):
    res = HttpResponse()
    try:
        mainClassList = []
        if request.user.has_perm('userlogin.NutrilliteArticleManage'):
            mainClassList.append("營養")
        if request.user.has_perm('userlogin.ArtistryArticleManage'):
            mainClassList.append("美容")
        if request.user.has_perm('userlogin.TechArticleManage'):
            mainClassList.append("科技")
        if request.user.has_perm('userlogin.AmwayQueenArticleManage'):
            mainClassList.append("金鍋")
        if request.user.has_perm('userlogin.OtherArticleManage'):
            mainClassList.append("其他")
        if request.user.has_perm('userlogin.ChainyenArticleManage'):
            mainClassList.append("總部會議/活動")
        if request.user.has_perm('userlogin.SpeechArticleManage'):
            mainClassList.append("演講廳")

        fileDatas = fileDataInfo.objects.filter(mainClass__mainClassName__in=mainClassList)
        data = fileDatas.get(id = articleID)

        keywordResult = fileDataKeywords.objects.filter(fileDataInfoID = data)
        keywordList = []
        for item in keywordResult:
            keywordList.append(item.keyword)
        tmp = {
            "id"       : data.id,
            "title"    : data.title,
            "DBClassCode"  : data.DBClass.DBClassCode,
            "mainClass": data.mainClass.mainClassName,
            "secClass" : data.secClass.secClassName,
            "keyword"  : "#" + '#'.join(keywordList),
            "describe" : data.describe,
            "point"    : data.point,
            "visible"  : data.visible
        }
        res.content = json.dumps(tmp)
        res.status_code = 200
    except:
        res.status_code = 503
    return res

@permission_required('userlogin.seeManagerArticlePage', login_url='/accounts/userlogin/')
def updateFileDataInfo(request, reportID = None):
    res = HttpResponse()
    try:
        modifier = request.user.user
        lastModify = str(datetime.datetime.now())
        id = request.POST['id']
        title = request.POST['title']
        mainClass = mainClassInfo.objects.get(mainClassName = request.POST["mainClass"]) 
        secClass = secClassInfo.objects.get(secClassName = request.POST['secClass'])
        keywordString = request.POST['keyword']
        describe = request.POST['describe']
        point = None
        if "point" in request.POST :
            point = int(request.POST['point'])
        visible = request.POST['visible']
        if visible == "true":
            visible = True
        else:
            visible = False
        fileData = fileDataInfo.objects.filter(id=id)
        if point and request.user.has_perm('userlogin.articlePointManage'):
            fileData.update(title=title, mainClass=mainClass,
                            secClass=secClass, describe=describe, point=point,
                            visible=visible, lastModify = lastModify)
        else:
            fileData.update(title=title, mainClass=mainClass,
                            secClass=secClass, describe=describe,visible=visible,
                            lastModify=lastModify)
        fileDataID = fileData.first()

        if "＃" in keywordString:
            keywordString = keywordString.replace("＃", "#")
        keywords = keywordString.split("#")[1:]
        if keywords:
            fileDataKeywords.objects.filter(fileDataInfoID = fileDataID).delete()
        for keyword in keywords:
            fDataKeyword = fileDataKeywords(fileDataInfoID = fileDataID, keyword = keyword)
            fDataKeyword.save()

        point = fileDataID.point

        fHistory = articleModifyHistory(fileDataID=fileDataID,modifier=modifier, title=title,
                                        secClass=secClass,keyword=keywordString,describe=describe,point=point,
                                        mainClass=mainClass,visible=visible, recordDate=lastModify)
        fHistory.save()

        if "reportID" in request.POST:
            handler = UserAccountInfo.objects.get(username = request.user.username)
            reportID = request.POST["reportID"]
            report = articleReport.objects.get(id =reportID)
            report.status = "finish"
            report.handler = handler
            report.handleDate = str(datetime.datetime.now())
            report.save()

        res.content = json.dumps({
            "id"       : id,
            "title"    : title,
            "mainClass": mainClass.mainClassName,
            "secClass" : secClass.secClassName,
            "describe" : describe,
            "point"    : point,
            "visible"  : visible
        })
        res.status_code = 200
        
    except:
        res.status_code = 503
    return res

@permission_required('userlogin.seeManagerArticlePage', login_url='/accounts/userlogin/')
def getArticleHistory(request):
    res = HttpResponse()
    try:
        id = request.POST['id']
        fileDataID = fileDataInfo.objects.get(id=id)
        aHistory = articleModifyHistory.objects.filter(fileDataID = fileDataID).order_by('-recordDate')
        
        historyList = []
        for data in aHistory:
            tmp = {
                "modifier":   data.modifier,
                "recordDate": str(data.recordDate),
                "title":      data.title,
                "mainClass":  data.mainClass.mainClassName,
                "secClass":   data.secClass.secClassName,
                "keyword":    data.keyword,
                "describe":   data.describe,
                "point":      data.point,
                "visible":    data.visible
            }
            historyList.append(tmp)
        res.status_code = 200
        res.content =  json.dumps(historyList)
    except:
        res.status_code = 503

    return res

@permission_required('userlogin.articleReport', login_url='/accounts/userlogin/')
def reportArticle(request):
    res = HttpResponse()
    try:
        reporter  = request.user.username
        articleID = request.POST["articleID"]
        reason = request.POST["reason"]
        recordDate = str(datetime.datetime.now())
        fileData = fileDataInfo.objects.get(id=articleID)
        userAccount = UserAccountInfo.objects.get(username=reporter)
        rReport = articleReport(reporter=userAccount, fileData=fileData,
                        reason=reason, status="report", recordDate=recordDate)
        rReport.save()
        res.status_code = 200
    except:
        res.status_code = 503
    return res

@permission_required('userlogin.seeManagerArticleReportPage', login_url='/accounts/userlogin/')
def removeArticleReport(request, reportID):
    res = HttpResponse()
    try:
        discardReason = ''
        if 'discardReason' in request.POST:
            discardReason = request.POST['discardReason']

        handler = UserAccountInfo.objects.get(username=request.user.username)
        rReport = articleReport.objects.get(id=int(reportID))
        rReport.status = "discard"
        rReport.handler = handler
        rReport.discardReason = discardReason
        rReport.handleDate = str(datetime.datetime.now())
        rReport.save()
        res.status_code = 200
    except:
        res.status_code = 503
    return res

@permission_required('userlogin.seeManagerArticleReportPage', login_url='/accounts/userlogin/')
def getArticleReport(request, status):
    res = HttpResponse()
    try:
        mainClassList = []
        if request.user.has_perm('userlogin.NutrilliteArticleManage'):
            mainClassList.append("營養")
        if request.user.has_perm('userlogin.ArtistryArticleManage'):
            mainClassList.append("美容")
        if request.user.has_perm('userlogin.TechArticleManage'):
            mainClassList.append("科技")
        if request.user.has_perm('userlogin.AmwayQueenArticleManage'):
            mainClassList.append("金鍋")
        if request.user.has_perm('userlogin.OtherArticleManage'):
            mainClassList.append("其他")
        if request.user.has_perm('userlogin.ChainyenArticleManage'):
            mainClassList.append("總部會議/活動")
        if request.user.has_perm('userlogin.SpeechArticleManage'):
            mainClassList.append("演講廳")

        articleDatas = articleReport.objects.filter(fileData__mainClass__mainClassName__in=mainClassList)
        articleDatas = articleDatas.filter(status=status)
        articleDataSummary = []
        for data in articleDatas:
            handler = ""
            if data.handler:
                handler = data.handler.user
            tmp = {
                "id"         : data.id,
                "articleID"  : data.fileData.id,
                "reporter"   : data.reporter.user,
                "handler"    : handler,
                "title"      : data.fileData.title,
                "mainClass"  : data.fileData.mainClass.mainClassName,
                "reason"     : data.reason,
                "status"     : data.status,
                "discardReason"     : data.discardReason,
                "recordDate" : str(data.recordDate),
                "handleDate"   : str(data.handleDate)
            }
            articleDataSummary.append(tmp)
        res.content = json.dumps(articleDataSummary)
        res.status_code = 200
    except:
        res.status_code = 503
    return res