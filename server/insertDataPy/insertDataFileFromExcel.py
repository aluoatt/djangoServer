
# -*- coding: utf-8 -*-
from userlogin.models import registerDDandDimInfo, amwayAwardInfo, chainYenClassInfo
from NutriliteSearchPage.models import fileDataInfo, mainClassInfo, secClassInfo, DBClassInfo , fileTypeInfo, sourceFromInfo,fileDataKeywords
from openpyxl import load_workbook
import re
from urllib.parse import urlparse

from datetime import datetime
def formatedDatetime(timeString):
    result = ""
    try:
        result = datetime.strptime(timeString, "%Y-%m-%d %H:%M:%S")
        return result
    except:
        result = ""
    try:
        result = datetime.strptime(timeString, "%Y-%m-%d %H:%M:%S.%f")
        return result
    except:
        result = ""
    try:
        result = datetime.strptime(timeString, "%m/%d/%Y")
        return result
    except:
        result = ""
    try:
        result = datetime.strptime(timeString, "%Y/%m/%d")
        return result
    except:
        result = ""
    try:
        result = datetime.strptime(timeString, "%Y/%m")
        return result
    except:
        result = ""
    try:
        result = datetime.strptime(timeString, "%Y%m")
        return result
    except:
        result = ""
    try:
        result = datetime.strptime(timeString, "%Y")
        return result
    except:
        result = ""
    return result

wb = load_workbook('insertDataPy/fileData20210908.xlsx')
sheet = wb.active
dataResult = []
init = True
pattern = re.compile(r'\s+')
rowCounter = 0
for i in range(2,sheet.max_row):
    rowIndex = str(i)
    # Get Original Value
    title = str(sheet['BC' + rowIndex].value or '')
    main  = str(sheet['K'  + rowIndex].value or '')
    sec   = str(sheet['N'  + rowIndex].value or '') + str(sheet['T' + rowIndex].value or '') + str(sheet['Z' + rowIndex].value or '') + str(sheet['AF' + rowIndex].value or '') + str(sheet['AL' + rowIndex].value or '') + str(sheet['AR'  + rowIndex].value or '') + str(sheet['AX'  + rowIndex].value or '')
    describe   = str(sheet['P' + rowIndex].value or '') + str(sheet['V' + rowIndex].value or '') + str(sheet['AB' + rowIndex].value or '') + str(sheet['AH' + rowIndex].value or '') + str(sheet['AN' + rowIndex].value or '') + str(sheet['AT' + rowIndex].value or '') + str(sheet['AZ' + rowIndex].value or '')
    sourceFrom = str(sheet['G' + rowIndex].value or '') + str(sheet['H' + rowIndex].value or '')
    PDFPath    = str(sheet['Q' + rowIndex].value or '') + str(sheet['W' + rowIndex].value or '') + str(sheet['AC' + rowIndex].value or '') + str(sheet['AI' + rowIndex].value or '') + str(sheet['AO' + rowIndex].value or '') + str(sheet['AU' + rowIndex].value or '') + str(sheet['BA' + rowIndex].value or '')
    filePath   = str(sheet['L' + rowIndex].value or '') + str(sheet['R' + rowIndex].value or '') + str(sheet['X' + rowIndex].value or '') + str(sheet['AD' + rowIndex].value or '') + str(sheet['AJ' + rowIndex].value or '') + str(sheet['AP' + rowIndex].value or '') + str(sheet['AV' + rowIndex].value or '')
    fileType   = str(sheet['M' + rowIndex].value or '') + str(sheet['S' + rowIndex].value or '') + str(sheet['Y'  + rowIndex].value or '') + str(sheet['AE'  + rowIndex].value or '') + str(sheet['AK'  + rowIndex].value or '') + str(sheet['AQ'  + rowIndex].value or '') + str(sheet['AW'  + rowIndex].value or '')
    keywords   = str(sheet['O' + rowIndex].value or '') + str(sheet['U' + rowIndex].value or '') + str(sheet['AA'  + rowIndex].value or '') + str(sheet['AG'  + rowIndex].value or '') + str(sheet['AM'  + rowIndex].value or '') + str(sheet['AS'  + rowIndex].value or '') + str(sheet['AY'  + rowIndex].value or '')
    sourceURL  = str(sheet['I' + rowIndex].value or '')
    sourceScreenshot = str(sheet['J' + rowIndex].value or '')
    characterName    = str(sheet['C' + rowIndex].value or '')
    characterClass   = str(sheet['D' + rowIndex].value or '').replace("教室", "")
    characterDD      = str(sheet['E' + rowIndex].value or '')
    lastModify       = str(sheet['A' + rowIndex].value or '')
    occurrenceDate   = str(sheet['BB' + rowIndex].value or '')
    visible          = True
    needWaterMark    = True
    permissionsLevel  = 1
    DBClass          = str(sheet['B'+rowIndex].value or '')
    downloadAble     = True
    if DBClass == "（A）公司資源（包含安麗公司；YouTube；行動大學；培訓資料.....）" or sourceFrom == "行動大學" or sourceFrom == "公司資料" or sourceFrom == "YouTube": 
        point = 2
        needWaterMark = False
    elif sourceFrom == "個人製作" or sourceFrom == "教室製作":
        point = 4
    elif sourceFrom == "總部製作":
        point = 6
    else:
        #should not exists
        point = 999
    if PDFPath == "":
        PDFPath = filePath
    oMain = main
    oSec = sec
    oODate = occurrenceDate
    oLastModify = lastModify
    oKeyword = keywords
    #格式化數值,以 DB 為主
    if "其他" in main:
        main = "其他"
    elif "演講廳" in main:
        main = "演講廳"
    elif "總部會議" in main:
        main = "總部會議/活動"
    try:

        PDFPath = urlparse(PDFPath).query.split("=")[1]
        filePath = urlparse(filePath).query.split("=")[1]
        main = mainClassInfo.objects.get(mainClassName = main)
        sec = secClassInfo.objects.get(secClassName = sec)
        fileType = fileTypeInfo.objects.get(fileTypeName = fileType)
        sourceFrom = sourceFromInfo.objects.get(sourceFromName = sourceFrom)
        DBClass = DBClassInfo.objects.get(DBClassName = DBClass)
        occurrenceDate = formatedDatetime(occurrenceDate)
        lastModify     = formatedDatetime(lastModify)
        if characterClass:
            characterClass = chainYenClassInfo.objects.get(ClassRoomName = characterClass)
        else:
            characterClass = None
        if not occurrenceDate:
            occurrenceDate = lastModify
        fileData = fileDataInfo(title = title, mainClass = main, secClass = sec,
                                describe = describe, PDF = PDFPath, file = filePath,
                                fileType = fileType, sourceFrom = sourceFrom,
                                sourceURL = sourceURL, sourceScreenshot = sourceScreenshot,
                                characterName = characterName, characterClass = characterClass,
                                characterDD = characterDD, occurrenceDate = occurrenceDate,
                                lastModify = lastModify, visible = visible, needWaterMark = needWaterMark,
                                point = point, permissionsLevel = permissionsLevel, DBClass = DBClass,
                                downloadAble = downloadAble)

        fileData.save()

        if "＃" in keywords:
            keywords = keywords.replace("＃", "#")
        keywords = keywords.split("#")
        for keyword in keywords:
            fDataKeyword = fileDataKeywords(fileDataInfoID = fileData, keyword = keyword)
            fDataKeyword.save()
    except Exception as e:
        print(e)
        print("occurrenceDate: ", oODate)
        print("lastModify: ", oLastModify)
        print("main: ", oMain)
        print("sec: ", oSec)
        print("describe: " + describe)
        print("keyword: " + oKeyword)
        print("===============================")
        if not oMain:
            break