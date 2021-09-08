# -*- coding: utf-8 -*-
from userlogin.models import registerDDandDimInfo, amwayAwardInfo
from openpyxl import load_workbook
import re

wb = load_workbook('insertDataPy/ChainYen20210831.xlsx')
sheet = wb.active
dataResult = []
init = True
pattern = re.compile(r'\s+')
for row in sheet.iter_rows(min_row=1, max_col=10, max_row=150, values_only=True):
    if init:
        init = False
        continue
    counter = 0
    endValue = False
    amwayAward = ""
    amwayNumber = ""
    amwayDiamond = ""
    main = ""
    sec = ""
    for value in row:
        if value:
            value = str(value)
            value = re.sub(pattern, '', value)
        if counter == 1 and value == None:
            endValue = True
            break
        if counter == 1:
            amwayNumber = value
        if counter == 2:
            amwayAward = value
        if counter == 3:
            main = value
        if counter == 4:
            if value:
                sec = value
        if counter == 5:
            amwayDiamond = value
        counter = counter + 1
    if endValue:
        break
    amAwardInfo = amwayAwardInfo.objects.get(amwayAward = amwayAward)
    r = registerDDandDimInfo(amwayAward = amAwardInfo, amwayNumber = amwayNumber,
                             amwayDiamond = amwayDiamond, main = main, sec = sec)
    r.save()
